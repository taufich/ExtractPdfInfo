import streamlit as st
import fitz  # PyMuPDF
import pandas as pd
import re
from collections import defaultdict
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import io

def extract_transcript_info(file):
    data = {
        "student_info": {},
        "courses_by_year": defaultdict(lambda: defaultdict(list)),
        "distribution_of_credit": {},
        "averages": {},
    }

    with fitz.open(stream=file.read(), filetype="pdf") as doc:
        full_text = "\n".join(page.get_text() for page in doc)

    student_info = re.search(
        r"Reg\. Nber\s+:\s+(\d+).*?Name\s+:\s+(.*?)\s+Major\s+:\s+(.*?)\nProgram\s+:\s+(.*?)\s+Minor\s+:",
        full_text, re.DOTALL
    )
    if student_info:
        data["student_info"] = {
            "Registration Number": student_info.group(1).strip(),
            "Name": student_info.group(2).strip(),
            "Major": student_info.group(3).strip(),
            "Program": student_info.group(4).strip()
        }

    academic_blocks = re.findall(r"Academic Year\s*:\s*(\d{4}-\d{4})(.*?)(?=Academic Year|Distribution of Credit)", full_text, re.DOTALL)
    for year, content in academic_blocks:
        semester_matches = re.findall(r"(\d)\s+([A-Z]+\s+\d+\s+.*?)\n(?=\d\s+[A-Z]+|\Z)", content, re.DOTALL)
        for semester, courses_block in semester_matches:
            course_lines = re.findall(r"([A-Z]{4}\s+\d+)\s+(.*?)\s+(\d+)\s+([\d.]+)\s+([√\*EC]*)", courses_block)
            for code, desc, credits, score, flags in course_lines:
                course = {
                    "Academic Year": year,
                    "Semester": f"Semester {semester}",
                    "Code": code.strip(),
                    "Description": desc.strip(),
                    "Credits": int(credits),
                    "Score": float(score),
                    "Flags": flags.strip()
                }
                data["courses_by_year"][year][f"Semester {semester}"].append(course)

    credit_match = re.search(r"Distribution of Credit\s*(.*?)Averages", full_text, re.DOTALL)
    if credit_match:
        for line in credit_match.group(1).strip().splitlines():
            if ":" in line:
                key, val = line.split(":")
                data["distribution_of_credit"][key.strip()] = int(val.strip())

    averages_match = re.search(r"Averages\s*(.*?)At Kigali", full_text, re.DOTALL)
    if averages_match:
        for line in averages_match.group(1).strip().splitlines():
            if ":" in line:
                key, val = line.split(":")
                data["averages"][key.strip()] = val.strip()

    return data

def generate_excel(data):
    rows = []
    for year, semesters in data["courses_by_year"].items():
        for semester, courses in semesters.items():
            for course in courses:
                rows.append(course)

    df_courses = pd.DataFrame(rows)
    df_info = pd.DataFrame([data["student_info"]])
    df_credit = pd.DataFrame([data["distribution_of_credit"]])
    df_avg = pd.DataFrame([data["averages"]])

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Student Info"
    for r in dataframe_to_rows(df_info, index=False, header=True):
        ws_info.append(r)

    ws_courses = wb.create_sheet("Courses")
    for r in dataframe_to_rows(df_courses, index=False, header=True):
        ws_courses.append(r)

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    for row in ws_courses.iter_rows(min_row=2, max_col=7):
        score = row[5].value
        flags = row[6].value
        if score is not None:
            if "EC" in flags or score < 12.0:
                for cell in row:
                    cell.fill = red_fill
            elif score >= 14.0:
                for cell in row:
                    cell.fill = green_fill

    for sheet_name, df in [("Credit Distribution", df_credit), ("Averages", df_avg)]:
        ws = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    # Save to in-memory file
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# Streamlit UI
st.set_page_config(page_title="Transcript Extractor", layout="wide")
st.title("📄 Transcript Extractor & Highlighter")

uploaded_pdf = st.file_uploader("Drag and drop your transcript PDF here", type="pdf")

if uploaded_pdf:
    transcript_data = extract_transcript_info(uploaded_pdf)

    st.subheader("👤 Student Info")
    st.json(transcript_data["student_info"])

    st.subheader("📚 Courses Summary")
    all_courses = []
    for year, semesters in transcript_data["courses_by_year"].items():
        for semester, courses in semesters.items():
            for course in courses:
                all_courses.append(course)
    df = pd.DataFrame(all_courses)
    st.dataframe(df, use_container_width=True)

    excel_data = generate_excel(transcript_data)
    st.download_button(
        label="📥 Download Excel with Highlights",
        data=excel_data,
        file_name="Transcript_Export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
